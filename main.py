#!/usr/bin/env python3

"""spreadsheet auto-fill

This script allows the user to automate the process to extract especific
information from a quotation to fill an .xlsx file.

This script is required to be located in the same directory as the .pdf
files as well as the main .xslx

This file contains the main function.

"""

import PyPDF2
import openpyxl
from pdf_regex_fun import pdf_regex
from recent_pdf import recent_pdf
from fill_excel import fill_excel
import sys
import ctypes

MB_OK = 0x0
ICON_EXLAIM=0x30
ICON_INFO = 0x40
MessageBox = ctypes.windll.user32.MessageBoxW

# Getting the most recent PDF file from the current directory
pdf_list = recent_pdf()
if len(pdf_list) == 0:
    MessageBox(
        None,
        'Sorry, inside this directory there are no PDF files',
        'No PDF file',
        MB_OK | ICON_EXLAIM
        )
    sys.exit(1)

# opening excel file and setting the worksheet
try:
    pipeline_file = openpyxl.load_workbook('Pipeline.xlsx')
    work_sheet = pipeline_file['Hoja1']
except FileNotFoundError:
    MessageBox(
        None,
        'The file Pipeline.xlsx doesn\'t exist inside this folder',
        'No .xlsx file',
        MB_OK | ICON_EXLAIM
        )
    sys.exit(1)
except:
    MessageBox(
        None,
        "Unexpected error:", sys.exc_info()[0],
        'No .xlsx file',
        MB_OK | ICON_EXLAIM
        )
    sys.exit(1)

# Getting the first empty row
empty_cell = work_sheet.max_row + 1

# Opening the PDF file and saving the text content in pdf_text variable
for recent_pdf_file in pdf_list:
    empty_cell = work_sheet.max_row + 1
    with open(recent_pdf_file, 'rb') as pdf_file:
        file_reader = PyPDF2.PdfFileReader(pdf_file)
        file_page = file_reader.getPage(0)
        pdf_text = file_page.extractText()
    if not pdf_text:
        MessageBox(
            None,
            "No data found in the file " + recent_pdf_file,
            'No data',
            MB_OK | ICON_EXLAIM
            )
        continue

# Obtaining the dictionary with the information extracted from pdf
    dict_values = pdf_regex(pdf_text, recent_pdf_file)
    if not dict_values:
        sys.exit(1)

    # Ideinfying if the quotation was already inserted in the .xlsx file
    i = 0
    for row in work_sheet.values:
        if (row[0] == dict_values['quot_number']):
            i += 1
    # Applying style and value to each cell from A# to Q#
    if (i == 0):
        fill_excel(dict_values, work_sheet, empty_cell)

pipeline_file.save('Pipeline.xlsx')
pipeline_file.close()

MessageBox(
        None,
        'Excel file updated',
        'Success',
        MB_OK | ICON_INFO
        )

# Make de proggram portable = "pip install pyinstaller",
# "got to direcotry -> app", "use pyinstaller --windowed
# --onefile --icon=./<icon.ico> <app_name>"
